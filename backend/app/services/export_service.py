"""
Excel / CSV export helpers for the admin panel.

Provides:
- build_csv(...) -> bytes
- build_xlsx(...) -> bytes
- export_response(...) -> FastAPI Response with the right content type
  and a Content-Disposition header so the browser downloads the file.

Used by the /export endpoints (orders, users, products, expenses).
"""

import csv
import io
from typing import Any, List, Optional, Sequence

from fastapi import Response

from app.core.logger import logger


def _stringify(value: Any) -> Any:
    """Return a JSON-safe / Excel-safe string representation of a value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (dict, list)):
        import json

        return json.dumps(value, ensure_ascii=False)
    return str(value)


def build_csv(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    """Build a CSV file (UTF-8 with BOM so Excel opens Bangla text correctly)."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow([_stringify(h) for h in headers])
    for row in rows:
        writer.writerow([_stringify(c) for c in row])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def build_xlsx(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    sheet_name: str = "Sheet1",
) -> bytes:
    """Build an .xlsx workbook with a styled header row."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover
        logger.error(f"❌ openpyxl not installed: {e}")
        raise RuntimeError("openpyxl is required for Excel export.")

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]  # Excel sheet-name limit

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_stringify(header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_stringify(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-fit column widths (capped for readability)
    for col_idx, _header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, min(len(rows) + 2, 200)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_response(
    data: bytes,
    filename: str,
    fmt: str = "xlsx",
) -> Response:
    """Wrap exported bytes into a downloadable FastAPI Response."""
    fmt = fmt.lower()
    if fmt == "csv":
        media_type = "text/csv; charset=utf-8"
    else:
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return Response(
        content=data,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )